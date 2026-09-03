"""
This script exists because five years of redemption history live only as
checkmarks on printed paper that got typed into a spreadsheet after the
fact. The new system needs one accurate starting point instead of trusting
staff to re-key hundreds of rows by hand, since that manual step is the
exact failure mode the club is trying to leave behind.

The old workbook organised data by calendar season (2021.2022, and so on).
The new app organises data by numeric range (1-10, 11-20, and so on)
instead. Those two ideas already line up one to one in the source file, a
season always sold exactly one ten slot range, so this script maps each
season directly onto its matching range rather than inventing a new
mapping.
"""

import json
import re
import sys
from collections import defaultdict

import openpyxl

SOURCE_FILE = sys.argv[1] if len(sys.argv) > 1 else "Whisky_Club_2025-26.xlsx"
OUTPUT_FILE = "seed.json"

# Seasons with no real activation date or consumption data (the season has
# not started selling yet) are skipped, since a fictional activation date
# for a range nobody has actually purchased would misrepresent expiry.
SEASONS_TO_IMPORT = [
    "2021.2022", "2022.2023", "2023.2024",
    "2024.2025", "2025.2026",
]

CURRENT_ROSTER_SEASON = "2026.2027"

SLOT_DEFINITION_PATTERN = re.compile(r"^\s*(\d+)\s*[-.]\s*(.*)$")
REMOVED_SUFFIX_PATTERN = re.compile(r"\(\s*[Rr]\s*\)\s*$")


def strip_removed_marker(raw_name):
    """
    A trailing (R), or the occasional autocorrected (R) symbol, is the only
    signal in the source data for "this person left the club that season".
    Splitting it out here means every downstream consumer works with a
    plain name and an explicit boolean instead of re parsing the string.
    """
    was_removed = bool(REMOVED_SUFFIX_PATTERN.search(raw_name)) or raw_name.strip().endswith("®")
    name = REMOVED_SUFFIX_PATTERN.sub("", raw_name).replace("®", "")
    name = re.sub(r"\s+", " ", name).strip()
    return name, was_removed


def season_start_year(season_label):
    """
    The club's program runs February to November, so February 1st of the
    season's first year is used as the fictional activation date for
    migrated data, matching when that season's range would have actually
    gone on sale.
    """
    return int(season_label.split(".")[0])


def range_id_for_slots(slot_numbers):
    """
    Every block in the source file is exactly ten consecutive numbers, so
    the range id is derived from the block's own minimum and maximum
    rather than trusted to match some separately maintained list.
    """
    valid_slots = [s for s in slot_numbers if s is not None]
    return f"{min(valid_slots)}-{max(valid_slots)}"


def read_season_blocks(worksheet):
    """
    A season's roster is not in a fixed range of rows, it grows or shrinks
    every year as members join and leave. Scanning for the header pattern
    (a blank name cell followed by ten slot numbers) is what lets this
    parser survive that instead of breaking when row counts shift.
    """
    rows = [
        [worksheet.cell(row=r, column=c).value for c in range(1, 12)]
        for r in range(1, worksheet.max_row + 1)
    ]

    blocks = []
    slot_names_found = {}

    row_index = 0
    while row_index < len(rows):
        row = rows[row_index]

        for name_column in (0, 3):
            value = row[name_column]
            if isinstance(value, str):
                match = SLOT_DEFINITION_PATTERN.match(value)
                if match:
                    slot_number = int(match.group(1))
                    slot_name = match.group(2).strip()
                    if slot_name:
                        slot_names_found[slot_number] = slot_name

        is_slot_header_row = row[0] is None and isinstance(row[1], int)
        if is_slot_header_row:
            slot_numbers = row[1:11]
            members = []
            member_row_index = row_index + 1
            while member_row_index < len(rows):
                member_row = rows[member_row_index]
                if member_row[0] is None:
                    break
                if isinstance(member_row[0], str) and member_row[0].startswith("Whisky Club"):
                    break
                name, was_removed = strip_removed_marker(member_row[0])
                consumed_flags = [member_row[k] == "ü" for k in range(1, 11)]
                members.append((name, was_removed, consumed_flags))
                member_row_index += 1
            blocks.append({"slotNumbers": slot_numbers, "members": members})
            row_index = member_row_index
        else:
            row_index += 1

    return blocks, slot_names_found


def build_seed_data(workbook):
    """
    Everything downstream (the app, the Apps Script backend, the Sheet
    mirror) expects one flat shape, so every imported season is folded
    into that shape here rather than leaking sheet layout into the rest
    of the codebase.
    """
    all_slot_names = {}
    member_ids_by_name = {}
    members = []
    range_memberships = []
    latest_status = {}

    def get_or_create_member_id(name):
        if name not in member_ids_by_name:
            member_id = f"m{len(members) + 1}"
            member_ids_by_name[name] = member_id
            members.append({"id": member_id, "name": name, "code": None, "active": False})
        return member_ids_by_name[name]

    for season_label in SEASONS_TO_IMPORT + [CURRENT_ROSTER_SEASON]:
        worksheet = workbook[season_label]
        blocks, slot_names_found = read_season_blocks(worksheet)
        all_slot_names.update(slot_names_found)

        if season_label == CURRENT_ROSTER_SEASON:
            # This season only supplies current membership status, it has
            # no real purchase to migrate since nobody has enrolled yet.
            for block in blocks:
                for name, was_removed, _ in block["members"]:
                    latest_status[name] = not was_removed
            continue

        activation_date = f"{season_start_year(season_label)}-02-01"

        for block in blocks:
            range_id = range_id_for_slots(block["slotNumbers"])
            for name, was_removed, consumed_flags in block["members"]:
                member_id = get_or_create_member_id(name)
                membership_id = f"rm_{member_id}_{range_id}"
                redemptions = [
                    {"slotNumber": slot, "consumed": consumed}
                    for slot, consumed in zip(block["slotNumbers"], consumed_flags)
                    if slot is not None
                ]
                range_memberships.append({
                    "id": membership_id,
                    "memberId": member_id,
                    "rangeId": range_id,
                    "activationDate": activation_date,
                    "paymentMethod": None,
                    "locked": False,
                    "redemptions": redemptions,
                })

    for member in members:
        member["active"] = latest_status.get(member["name"], False)

    whiskey_slots = [
        {"number": n, "name": all_slot_names.get(n)}
        for n in range(1, 101)
    ]

    return {"members": members, "whiskeySlots": whiskey_slots, "rangeMemberships": range_memberships}


def attach_member_codes(seed_data, workbook):
    """
    The Master List predates the current roster and uses slightly
    different name formatting, so codes are matched on a normalised name
    rather than assumed to line up row for row.
    """
    def normalise(name):
        return re.sub(r"\s+", " ", name).strip().lower()

    worksheet = workbook["Master List"]
    code_by_name = {}
    for r in range(2, worksheet.max_row + 1):
        name = worksheet.cell(row=r, column=1).value
        code = worksheet.cell(row=r, column=3).value
        if isinstance(name, str) and code:
            code_by_name[normalise(name)] = code

    for member in seed_data["members"]:
        member["code"] = code_by_name.get(normalise(member["name"]))


def use_codes_as_member_ids(seed_data):
    """
    The app now uses a member's code as their id, since that is what
    staff already look someone up by at the bar. Historical members were
    originally given generated ids (m1, m2, ...) before their code was
    known, so those ids are rewritten here, along with every membership
    that references them, once the real code has been attached.

    A member with no code on file (missing from the old Master List)
    keeps a readable fallback id instead, flagged clearly so it can be
    fixed by hand once a real code is known, rather than silently left
    as an opaque generated id.
    """
    old_id_to_new_id = {}
    for member in seed_data["members"]:
        old_id = member["id"]
        new_id = member["code"] if member["code"] else f"NOCODE_{old_id}"
        old_id_to_new_id[old_id] = new_id
        member["id"] = new_id

    for membership in seed_data["rangeMemberships"]:
        membership["memberId"] = old_id_to_new_id[membership["memberId"]]


if __name__ == "__main__":
    workbook = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
    seed_data = build_seed_data(workbook)
    attach_member_codes(seed_data, workbook)
    use_codes_as_member_ids(seed_data)

    active_count = sum(1 for m in seed_data["members"] if m["active"])
    missing_code_count = sum(1 for m in seed_data["members"] if m["id"].startswith("NOCODE_"))
    print(f"Members parsed: {len(seed_data['members'])}")
    print(f"Currently active: {active_count}")
    print(f"Range memberships parsed: {len(seed_data['rangeMemberships'])}")
    print(f"Members missing a code (need manual fix): {missing_code_count}")

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(seed_data, f, indent=2, ensure_ascii=False)
    print(f"Written to {OUTPUT_FILE}")
